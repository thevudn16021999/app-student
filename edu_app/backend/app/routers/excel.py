"""
Router: Import/Export Excel
"""
import io
import urllib.parse
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..database import get_db
from .. import crud, schemas, models

try:
    import openpyxl
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/api/excel", tags=["Import/Export"])


@router.post("/import/{classroom_id}")
async def import_excel(classroom_id: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import danh sách học sinh từ file Excel (.xlsx) hoặc CSV.
    Cột: Tên/Họ và tên (bắt buộc), Điểm (tùy chọn, mặc định 0)
    """
    if not file.filename.endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file .xlsx hoặc .csv")

    contents = await file.read()
    imported = []
    errors = []

    try:
        if file.filename.endswith('.csv'):
            # Parse CSV
            import csv
            text = contents.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            # Parse XLSX
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [cell.value for cell in ws[1] if cell.value]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                if any(v for v in row_dict.values()):
                    rows.append(row_dict)

        # Tìm cột tên
        for idx, row in enumerate(rows):
            name = None
            points = 0

            # Tìm tên học sinh
            for key in ['Tên', 'Họ và tên', 'Ho va ten', 'Name', 'ten', 'name']:
                if key in row and row[key]:
                    name = str(row[key]).strip()
                    break

            if not name:
                errors.append(f"Dòng {idx + 2}: Không tìm thấy tên học sinh")
                continue

            # Tìm điểm
            for key in ['Điểm', 'Diem', 'Points', 'Score', 'diem', 'points']:
                if key in row and row[key] is not None:
                    try:
                        points = int(float(str(row[key])))
                    except (ValueError, TypeError):
                        points = 0
                    break

            # Tạo học sinh
            student_data = schemas.StudentCreate(
                name=name,
                order_number=idx + 1,
                total_points=max(0, points)
            )
            student = crud.create_student(db, classroom_id, student_data)
            imported.append({"name": name, "points": points})

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi đọc file: {str(e)}")

    return {
        "imported": len(imported),
        "errors": errors,
        "students": imported
    }


@router.get("/export/{classroom_id}")
def export_excel(classroom_id: str, db: Session = Depends(get_db)):
    """
    Export dữ liệu lớp ra file Excel gồm 3 sheet:
    1. Danh sách: Tên | Điểm | Hạng | Tổng cộng | Tổng trừ
    2. Lịch sử: Tên | Thời gian | Thay đổi | Lý do | Điểm sau
    3. Quà đã đổi: Tên | Quà | Điểm tiêu | Thời gian
    """
    students = crud.get_students(db, classroom_id)
    classroom = db.query(models.Classroom).filter(models.Classroom.id == classroom_id).first()
    if not classroom:
        raise HTTPException(status_code=404, detail="Không tìm thấy lớp học")

    wb = openpyxl.Workbook()

    # ---------- Sheet 1: Danh sách ----------
    ws1 = wb.active
    ws1.title = "Danh sách"
    ws1.append(["STT", "Tên", "Điểm", "Hạng", "Tổng cộng", "Tổng trừ"])

    rank_map = {"bronze": "Đồng", "silver": "Bạc", "gold": "Vàng", "diamond": "Kim Cương"}

    for student in students:
        total_add = sum(h.change for h in student.point_history if h.change > 0)
        total_sub = sum(h.change for h in student.point_history if h.change < 0)
        ws1.append([
            student.order_number,
            student.name,
            student.total_points,
            rank_map.get(student.rank, student.rank),
            total_add,
            total_sub
        ])

    # ---------- Sheet 2: Lịch sử ----------
    ws2 = wb.create_sheet("Lịch sử")
    ws2.append(["Tên", "Thời gian", "Thay đổi", "Lý do", "Điểm sau"])

    for student in students:
        for h in sorted(student.point_history, key=lambda x: x.timestamp, reverse=True):
            ws2.append([
                student.name,
                h.timestamp.strftime("%d/%m/%Y %H:%M"),
                h.change,
                h.reason,
                h.points_after
            ])

    # ---------- Sheet 3: Quà đã đổi ----------
    ws3 = wb.create_sheet("Quà đã đổi")
    ws3.append(["Tên", "Quà", "Điểm tiêu", "Thời gian"])

    for student in students:
        for r in sorted(student.rewards_redeemed, key=lambda x: x.timestamp, reverse=True):
            ws3.append([
                student.name,
                r.reward_name,
                r.points_spent,
                r.timestamp.strftime("%d/%m/%Y %H:%M")
            ])

    # Ghi ra buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"xephang_{classroom.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    # Encode tên file để tránh lỗi latin-1 với ký tự tiếng Việt
    encoded_filename = urllib.parse.quote(filename)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )
